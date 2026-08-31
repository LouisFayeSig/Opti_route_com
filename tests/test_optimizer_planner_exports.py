from __future__ import annotations

import io
import re

import pandas as pd
from openpyxl import load_workbook

from opti_route.exporting import (
    csv_bytes,
    excel_bytes,
    google_maps_url,
    pdf_bytes,
    sanitize_spreadsheet_value,
)
from opti_route.optimizer import optimize_route
from opti_route.planner import StartPoint, build_route_plan


def test_optimizer_visits_each_node_and_returns() -> None:
    matrix = [
        [0, 10, 20, 30],
        [10, 0, 10, 20],
        [20, 10, 0, 10],
        [30, 20, 10, 0],
    ]
    route = optimize_route(matrix, matrix, return_to_start=True, time_limit_seconds=1)
    assert route[0] == route[-1] == 0
    assert sorted(route[1:-1]) == [1, 2, 3]


def test_planner_builds_route_and_exports() -> None:
    clients = pd.DataFrame(
        {
            "client_id": ["A", "B", "C"],
            "client_name": ["Alpha", "Beta", "Gamma"],
            "salesperson": ["Morgan"] * 3,
            "address": ["Adresse A", "Adresse B", "Adresse C"],
            "address_2": [pd.NA] * 3,
            "address_3": [pd.NA] * 3,
            "postal_code": ["14000", "14120", "14200"],
            "city": ["Caen", "Mondeville", "Hérouville-Saint-Clair"],
            "country": ["France"] * 3,
            "latitude": [49.183, 49.174, 49.205],
            "longitude": [-0.370, -0.320, -0.335],
            "full_address": ["A", "B", "C"],
        }
    )
    plan = build_route_plan(
        clients,
        StartPoint(49.1829, -0.3707, "Caen"),
        radius_km=20,
        max_visits=3,
        max_duration_hours=4,
        return_to_start=True,
        objective="time",
    )

    assert plan.visit_count == 3
    assert plan.route_coordinates[0] == plan.route_coordinates[-1]
    assert plan.total_distance_m > 0
    assert plan.itinerary_table()["Étape"].tolist() == ["Départ", "1", "2", "3", "Retour"]
    assert csv_bytes(plan).startswith(b"\xef\xbb\xbf")
    assert "Départ" in csv_bytes(plan).decode("utf-8-sig")
    assert excel_bytes(plan).startswith(b"PK")
    pdf = pdf_bytes(plan)
    assert pdf.startswith(b"%PDF")
    assert b"/Subtype /Image" in pdf
    assert re.fullmatch(r"tournee_commerciale_\d{8}_\d{6}", plan.export_stem)
    assert "google.com/maps/dir" in google_maps_url(plan)


def test_planner_keeps_a_specific_arrival_after_all_visits() -> None:
    clients = pd.DataFrame(
        {
            "client_id": ["A", "B"],
            "client_name": ["Alpha", "Beta"],
            "salesperson": ["Morgan", "Morgan"],
            "address": ["Adresse A", "Adresse B"],
            "address_2": [pd.NA, pd.NA],
            "address_3": [pd.NA, pd.NA],
            "postal_code": ["14000", "14120"],
            "city": ["Caen", "Mondeville"],
            "country": ["France", "France"],
            "latitude": [49.183, 49.174],
            "longitude": [-0.370, -0.320],
            "full_address": ["A", "B"],
        }
    )
    arrival = StartPoint(49.250, -0.250, "Agence")

    plan = build_route_plan(
        clients,
        StartPoint(49.1829, -0.3707, "Départ"),
        radius_km=20,
        max_visits=2,
        max_duration_hours=None,
        return_to_start=True,
        objective="time",
        end=arrival,
    )

    assert not plan.return_to_start
    assert plan.end == arrival
    assert plan.route_coordinates[-1] == (arrival.latitude, arrival.longitude)
    assert plan.itinerary_table().iloc[-1]["Étape"] == "Arrivée"
    assert plan.itinerary_table().iloc[-1]["Client"] == "Agence"


def test_planner_prioritizes_the_closest_selected_companies() -> None:
    clients = pd.DataFrame(
        {
            "client_id": ["FAR", "NEAR"],
            "client_name": ["Entreprise éloignée", "Entreprise proche"],
            "salesperson": ["Morgan", "Morgan"],
            "address": ["Adresse éloignée", "Adresse proche"],
            "address_2": [pd.NA, pd.NA],
            "address_3": [pd.NA, pd.NA],
            "postal_code": ["75000", "14000"],
            "city": ["Paris", "Caen"],
            "country": ["France", "France"],
            "latitude": [48.8566, 49.184],
            "longitude": [2.3522, -0.371],
            "full_address": ["Paris", "Caen"],
        }
    )

    plan = build_route_plan(
        clients,
        StartPoint(49.1829, -0.3707, "Départ"),
        radius_km=300,
        max_visits=1,
        max_duration_hours=None,
        return_to_start=False,
        objective="time",
    )

    assert plan.visit_count == 1
    assert plan.table.iloc[0]["Client"] == "Entreprise proche"


def test_spreadsheet_formula_prefixes_are_escaped() -> None:
    dangerous_values = [
        '=HYPERLINK("https://example.test")',
        "+1+1",
        "-1+1",
        "@SUM(1,1)",
        "  =SUM(1,1)",
        "＝SUM(1,1)",
    ]

    for value in dangerous_values:
        assert sanitize_spreadsheet_value(value) == "'" + value
    assert sanitize_spreadsheet_value("Client normal") == "Client normal"
    assert sanitize_spreadsheet_value(-12.5) == -12.5


def test_csv_and_excel_exports_keep_untrusted_values_as_text() -> None:
    clients = pd.DataFrame(
        {
            "client_id": ["A"],
            "client_name": ['=HYPERLINK("https://example.test")'],
            "salesperson": ["Morgan"],
            "address": ["Adresse A"],
            "address_2": [pd.NA],
            "address_3": [pd.NA],
            "postal_code": ["14000"],
            "city": ["@SUM(1,1)"],
            "country": ["France"],
            "latitude": [49.183],
            "longitude": [-0.370],
            "full_address": ["+Adresse A"],
        }
    )
    plan = build_route_plan(
        clients,
        StartPoint(49.1829, -0.3707, "=Point de départ"),
        radius_km=20,
        max_visits=1,
        max_duration_hours=4,
        return_to_start=False,
        objective="time",
    )

    csv_text = csv_bytes(plan).decode("utf-8-sig")
    assert "'=HYPERLINK" in csv_text
    assert "'@SUM" in csv_text
    assert "'+Adresse A" in csv_text

    workbook = load_workbook(io.BytesIO(excel_bytes(plan)), data_only=False)
    route_sheet = workbook["Tournée"]
    headers = [cell.value for cell in route_sheet[1]]
    client_column = headers.index("Client") + 1
    city_column = headers.index("Ville") + 1
    address_column = headers.index("Adresse") + 1
    client_cell = route_sheet.cell(row=3, column=client_column)
    city_cell = route_sheet.cell(row=3, column=city_column)
    address_cell = route_sheet.cell(row=3, column=address_column)
    assert client_cell.value.startswith("'=HYPERLINK")
    assert city_cell.value.startswith("'@SUM")
    assert address_cell.value.startswith("'+Adresse")
    assert client_cell.data_type == city_cell.data_type == address_cell.data_type == "s"

    summary_sheet = workbook["Synthèse"]
    assert summary_sheet["B2"].value == "'=Point de départ"
    assert summary_sheet["B2"].data_type == "s"
